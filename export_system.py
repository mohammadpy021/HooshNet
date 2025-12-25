"""
Data Export System for HooshNet VPN Bot
Exports users, orders, and payments to CSV/Excel formats
"""

import csv
import io
import logging
from typing import Optional, Dict, List
from datetime import datetime
from enum import Enum

logger = logging.getLogger(__name__)


class ExportFormat(Enum):
    """Export file formats"""
    CSV = 'csv'
    EXCEL = 'xlsx'  # Requires openpyxl


class ExportType(Enum):
    """Types of data to export"""
    USERS = 'users'
    ORDERS = 'orders'
    PAYMENTS = 'payments'
    SERVICES = 'services'
    INVOICES = 'invoices'


class DataExporter:
    """Exports data to various formats"""
    
    def __init__(self, db=None):
        self.db = db
    
    def set_database(self, db):
        """Set database instance"""
        self.db = db
    
    # ==================== User Export ====================
    
    def export_users(self, format: ExportFormat = ExportFormat.CSV, 
                     filters: Dict = None) -> Optional[bytes]:
        """
        Export users to CSV/Excel
        
        Args:
            format: Export format (CSV or Excel)
            filters: Optional filters (is_admin, is_active, date_range, etc.)
            
        Returns:
            Bytes of the exported file
        """
        if not self.db:
            return None
        
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor(dictionary=True)
                
                query = '''
                    SELECT 
                        u.id,
                        u.telegram_id,
                        u.username,
                        u.first_name,
                        u.last_name,
                        u.balance,
                        u.is_admin,
                        u.admin_role,
                        u.is_active,
                        u.is_banned,
                        u.referred_by,
                        u.referral_code,
                        u.total_referrals,
                        u.total_referral_earnings,
                        u.created_at,
                        u.last_activity,
                        u.total_spent,
                        u.total_services
                    FROM users u
                    WHERE 1=1
                '''
                
                params = []
                if filters:
                    if filters.get('is_admin'):
                        query += " AND u.is_admin = 1"
                    if filters.get('is_active') is not None:
                        query += " AND u.is_active = %s"
                        params.append(1 if filters['is_active'] else 0)
                    if filters.get('date_from'):
                        query += " AND u.created_at >= %s"
                        params.append(filters['date_from'])
                    if filters.get('date_to'):
                        query += " AND u.created_at <= %s"
                        params.append(filters['date_to'])
                
                query += " ORDER BY u.created_at DESC"
                
                cursor.execute(query, tuple(params))
                users = cursor.fetchall()
                
                # Define headers
                headers = [
                    'ID', 'Telegram ID', 'Username', 'First Name', 'Last Name',
                    'Balance', 'Is Admin', 'Admin Role', 'Is Active', 'Is Banned',
                    'Referred By', 'Referral Code', 'Total Referrals', 'Referral Earnings',
                    'Created At', 'Last Activity', 'Total Spent', 'Total Services'
                ]
                
                return self._export_to_format(users, headers, format, 'users')
                
        except Exception as e:
            logger.error(f"Error exporting users: {e}")
            return None
    
    # ==================== Orders Export ====================
    
    def export_orders(self, format: ExportFormat = ExportFormat.CSV,
                      filters: Dict = None) -> Optional[bytes]:
        """Export orders/invoices to CSV/Excel"""
        if not self.db:
            return None
        
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor(dictionary=True)
                
                query = '''
                    SELECT 
                        i.id,
                        i.order_id,
                        u.telegram_id,
                        u.username,
                        p.name as panel_name,
                        i.gb_amount,
                        i.duration_days,
                        i.amount,
                        i.original_amount,
                        i.discount_amount,
                        i.status,
                        i.payment_method,
                        i.purchase_type,
                        i.created_at,
                        i.paid_at
                    FROM invoices i
                    JOIN users u ON i.user_id = u.id
                    LEFT JOIN panels p ON i.panel_id = p.id
                    WHERE 1=1
                '''
                
                params = []
                if filters:
                    if filters.get('status'):
                        query += " AND i.status = %s"
                        params.append(filters['status'])
                    if filters.get('date_from'):
                        query += " AND i.created_at >= %s"
                        params.append(filters['date_from'])
                    if filters.get('date_to'):
                        query += " AND i.created_at <= %s"
                        params.append(filters['date_to'])
                
                query += " ORDER BY i.created_at DESC"
                
                cursor.execute(query, tuple(params))
                orders = cursor.fetchall()
                
                headers = [
                    'ID', 'Order ID', 'Telegram ID', 'Username', 'Panel',
                    'GB Amount', 'Duration Days', 'Amount', 'Original Amount',
                    'Discount', 'Status', 'Payment Method', 'Purchase Type',
                    'Created At', 'Paid At'
                ]
                
                return self._export_to_format(orders, headers, format, 'orders')
                
        except Exception as e:
            logger.error(f"Error exporting orders: {e}")
            return None
    
    # ==================== Payments Export ====================
    
    def export_payments(self, format: ExportFormat = ExportFormat.CSV,
                        filters: Dict = None) -> Optional[bytes]:
        """Export payment transactions to CSV/Excel"""
        if not self.db:
            return None
        
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor(dictionary=True)
                
                query = '''
                    SELECT 
                        bt.id,
                        u.telegram_id,
                        u.username,
                        bt.amount,
                        bt.transaction_type,
                        bt.description,
                        bt.reference_id,
                        bt.created_at
                    FROM balance_transactions bt
                    JOIN users u ON bt.user_id = u.id
                    WHERE 1=1
                '''
                
                params = []
                if filters:
                    if filters.get('transaction_type'):
                        query += " AND bt.transaction_type = %s"
                        params.append(filters['transaction_type'])
                    if filters.get('date_from'):
                        query += " AND bt.created_at >= %s"
                        params.append(filters['date_from'])
                    if filters.get('date_to'):
                        query += " AND bt.created_at <= %s"
                        params.append(filters['date_to'])
                
                query += " ORDER BY bt.created_at DESC"
                
                cursor.execute(query, tuple(params))
                payments = cursor.fetchall()
                
                headers = [
                    'ID', 'Telegram ID', 'Username', 'Amount', 'Type',
                    'Description', 'Reference ID', 'Created At'
                ]
                
                return self._export_to_format(payments, headers, format, 'payments')
                
        except Exception as e:
            logger.error(f"Error exporting payments: {e}")
            return None
    
    # ==================== Services Export ====================
    
    def export_services(self, format: ExportFormat = ExportFormat.CSV,
                        filters: Dict = None) -> Optional[bytes]:
        """Export client services to CSV/Excel"""
        if not self.db:
            return None
        
        try:
            with self.db.get_connection() as conn:
                cursor = conn.cursor(dictionary=True)
                
                query = '''
                    SELECT 
                        c.id,
                        c.client_name,
                        c.client_uuid,
                        u.telegram_id,
                        u.username,
                        p.name as panel_name,
                        c.protocol,
                        c.total_gb,
                        c.used_gb,
                        c.expire_days,
                        c.is_active,
                        c.status,
                        c.created_at,
                        c.expires_at,
                        c.last_used
                    FROM clients c
                    JOIN users u ON c.user_id = u.id
                    JOIN panels p ON c.panel_id = p.id
                    WHERE 1=1
                '''
                
                params = []
                if filters:
                    if filters.get('status'):
                        query += " AND c.status = %s"
                        params.append(filters['status'])
                    if filters.get('panel_id'):
                        query += " AND c.panel_id = %s"
                        params.append(filters['panel_id'])
                    if filters.get('is_active') is not None:
                        query += " AND c.is_active = %s"
                        params.append(1 if filters['is_active'] else 0)
                
                query += " ORDER BY c.created_at DESC"
                
                cursor.execute(query, tuple(params))
                services = cursor.fetchall()
                
                headers = [
                    'ID', 'Client Name', 'UUID', 'Telegram ID', 'Username',
                    'Panel', 'Protocol', 'Total GB', 'Used GB', 'Expire Days',
                    'Is Active', 'Status', 'Created At', 'Expires At', 'Last Used'
                ]
                
                return self._export_to_format(services, headers, format, 'services')
                
        except Exception as e:
            logger.error(f"Error exporting services: {e}")
            return None
    
    # ==================== Format Helpers ====================
    
    def _export_to_format(self, data: List[Dict], headers: List[str], 
                          format: ExportFormat, filename: str) -> Optional[bytes]:
        """Export data to specified format"""
        if format == ExportFormat.CSV:
            return self._to_csv(data, headers)
        elif format == ExportFormat.EXCEL:
            return self._to_excel(data, headers, filename)
        else:
            return self._to_csv(data, headers)
    
    def _to_csv(self, data: List[Dict], headers: List[str]) -> bytes:
        """Convert data to CSV bytes"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow(headers)
        
        # Write data
        for row in data:
            values = []
            for key in row.keys():
                value = row[key]
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                values.append(str(value) if value is not None else '')
            writer.writerow(values)
        
        return output.getvalue().encode('utf-8-sig')  # BOM for Excel compatibility
    
    def _to_excel(self, data: List[Dict], headers: List[str], 
                  filename: str) -> Optional[bytes]:
        """Convert data to Excel bytes"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = filename
            
            # Header styling
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row_idx, row in enumerate(data, 2):
                for col_idx, key in enumerate(row.keys(), 1):
                    value = row[key]
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-fit columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
            
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return self._to_csv(data, headers)
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            return self._to_csv(data, headers)
    
    def get_export_filename(self, export_type: ExportType, format: ExportFormat) -> str:
        """Generate filename for export"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        extension = 'csv' if format == ExportFormat.CSV else 'xlsx'
        return f"hooshnet_{export_type.value}_{timestamp}.{extension}"


# Global instance
data_exporter = DataExporter()
